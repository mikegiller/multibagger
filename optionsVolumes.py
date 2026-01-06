import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from datetime import datetime

def main():
    # Prompt for ticker
    ticker = input("Enter a stock ticker (e.g., AAPL): ").upper().strip()
    print(f"\nFetching option data for {ticker}...")

    try:
        stock = yf.Ticker(ticker)
        expirations = stock.options

        if not expirations:
            print("No options data available for this ticker.")
            return

        call_data = []
        put_data = []

        for exp in expirations:
            try:
                chain = stock.option_chain(exp)
                calls_volume = chain.calls['volume'].fillna(0).sum()
                puts_volume = chain.puts['volume'].fillna(0).sum()

                call_data.append({"Expiration Date": exp, "Calls Volume": calls_volume})
                put_data.append({"Expiration Date": exp, "Puts Volume": puts_volume})

            except Exception as e:
                print(f"Skipping {exp} due to error: {e}")

        # Convert to DataFrames
        df_calls = pd.DataFrame(call_data)
        df_puts = pd.DataFrame(put_data)

        for df in [df_calls, df_puts]:
            df["Expiration Date"] = pd.to_datetime(df["Expiration Date"])
            df.sort_values("Expiration Date", inplace=True)

        # Create workbook
        filename = f"{ticker}_options_volumes_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb = Workbook()

        # --- Calls Sheet ---
        ws_calls = wb.active
        ws_calls.title = "Calls Volume"
        ws_calls.append(["Expiration Date", "Calls Volume"])
        for row in df_calls.itertuples(index=False):
            ws_calls.append([row[0].strftime('%Y-%m-%d'), row[1]])

        # Add chart for Calls
        chart_calls = LineChart()
        chart_calls.title = f"{ticker} Calls Volume by Expiration Date"
        chart_calls.y_axis.title = "Calls Volume"
        chart_calls.x_axis.title = "Expiration Date"
        data_ref = Reference(ws_calls, min_col=2, min_row=1, max_row=ws_calls.max_row)
        cats_ref = Reference(ws_calls, min_col=1, min_row=2, max_row=ws_calls.max_row)
        chart_calls.add_data(data_ref, titles_from_data=True)
        chart_calls.set_categories(cats_ref)
        ws_calls.add_chart(chart_calls, "E5")

        # --- Puts Sheet ---
        ws_puts = wb.create_sheet("Puts Volume")
        ws_puts.append(["Expiration Date", "Puts Volume"])
        for row in df_puts.itertuples(index=False):
            ws_puts.append([row[0].strftime('%Y-%m-%d'), row[1]])

        # Add chart for Puts
        chart_puts = LineChart()
        chart_puts.title = f"{ticker} Puts Volume by Expiration Date"
        chart_puts.y_axis.title = "Puts Volume"
        chart_puts.x_axis.title = "Expiration Date"
        data_ref2 = Reference(ws_puts, min_col=2, min_row=1, max_row=ws_puts.max_row)
        cats_ref2 = Reference(ws_puts, min_col=1, min_row=2, max_row=ws_puts.max_row)
        chart_puts.add_data(data_ref2, titles_from_data=True)
        chart_puts.set_categories(cats_ref2)
        ws_puts.add_chart(chart_puts, "E5")

        # Save workbook
        wb.save(filename)
        print(f"\n✅ Excel file created: {filename}")

    except Exception as e:
        print(f"Error fetching data for {ticker}: {e}")

if __name__ == "__main__":
    main()
