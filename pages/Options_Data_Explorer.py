# app.py
import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import yfinance as yf
from io import BytesIO
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import plotly.graph_objects as go

# SSL fix for Mac
import ssl
import certifi
ssl._create_default_https_context = ssl._create_unverified_context

# Optional: Google Gemini (only needed for AI analysis)
try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

# Current date as per user instruction
TODAY = datetime(2026, 1, 6)

# --- Page settings ---
st.set_page_config(page_title="Call Options Viewer", layout="wide")
st.title("📈 Call Options Viewer")
st.write("Enter a ticker to view call options, toggle columns, and see return potential graph (positive only).")

# --- Gemini API Key Setup (in sidebar) ──────────────────────────────
with st.sidebar:
    st.header("🤖 AI Analysis Settings")
    
    if not GEMINI_AVAILABLE:
        st.warning("⚠️ Google Gemini not installed")
        st.code("pip install google-generativeai", language="bash")
        gemini_api_key = None
    else:
        gemini_api_key = st.text_input(
            "Gemini API Key", 
            type="password",
            help="Get free API key at: https://aistudio.google.com/app/apikey",
            value=""
        )
        
        if gemini_api_key:
            try:
                genai.configure(api_key=gemini_api_key)
                st.success("✅ Gemini API configured")
            except Exception as e:
                st.error(f"❌ API configuration failed: {str(e)}")
        else:
            st.info("💡 Add API key to enable AI analysis")
        
        st.markdown("---")
        st.caption("**Free Tier**: 1,500 requests/day")
        st.caption("[Get API Key →](https://aistudio.google.com/app/apikey)")

# --- Session state ---
for key in ["ticker_input", "expiry", "calls_data", "summary_data", "column_names"]:
    if key not in st.session_state:
        st.session_state[key] = None

# --- Reset button ---
if st.button("🔄 Reset"):
    for key in ["ticker_input", "expiry", "calls_data", "summary_data", "column_names"]:
        st.session_state[key] = None

# --- Ticker input ---
ticker_input = st.text_input("Enter ticker symbol (required):", value=st.session_state.ticker_input or "").upper()
st.session_state.ticker_input = ticker_input

# --- Fetch expirations ---
expirations = []
if ticker_input:
    try:
        ticker = yf.Ticker(ticker_input)
        hist = ticker.history(period="1d")
        if hist.empty:
            st.error("No data found for ticker.")
        elif not ticker.options:
            st.error("No options available for this ticker.")
        else:
            expirations = ticker.options
    except Exception as e:
        st.error(f"Error fetching ticker info: {e}")

# --- Expiration dropdown with days to expiry ---
expiry = None
if expirations:
    # Sort descending (soonest expiration first)
    expirations_sorted = sorted(expirations, reverse=True)
    
    # Format options with days remaining
    expiry_options = ["Select expiration date..."]
    for exp_date_str in expirations_sorted:
        exp_date = datetime.strptime(exp_date_str, "%Y-%m-%d")
        days_left = (exp_date - TODAY).days
        label = f"{exp_date_str} ({days_left} days to expiration)"
        expiry_options.append(label)
    
    # Find index of previously selected expiry
    selected_index = 0
    if st.session_state.expiry:
        prev_label = f"{st.session_state.expiry} ({(datetime.strptime(st.session_state.expiry, '%Y-%m-%d') - TODAY).days} days to expiration)"
        if prev_label in expiry_options:
            selected_index = expiry_options.index(prev_label)

    selected_label = st.selectbox(
        "Select expiration date (oldest first):",
        expiry_options,
        index=selected_index
    )
    
    # Extract the date part back from selected label
    if selected_label != "Select expiration date...":
        expiry = selected_label.split(" ")[0]  # e.g. "2026-01-16"
    else:
        expiry = None
    
    st.session_state.expiry = expiry

# --- Data Refresh button ---
refresh_pressed = st.button("🔄 Data Refresh")

# --- Cached data fetch function ---
@st.cache_data(ttl=600, show_spinner="Fetching fresh options data from Yahoo Finance...")
def fetch_option_data_cached(ticker_input, expiry):
    ticker = yf.Ticker(ticker_input)
    hist = ticker.history(period="1d")
    last_price = round(hist["Close"].iloc[-1], 2) if not hist.empty else None
    expiry_date = datetime.strptime(expiry, "%Y-%m-%d")
    days_to_expiry = (expiry_date - TODAY).days
    years_to_expiry = days_to_expiry / 365 if days_to_expiry > 0 else 0.0001

    summary_df = pd.DataFrame({
        "Ticker": [ticker_input],
        "Last Price": [last_price],
        "Expiration": [expiry],
        "Days to Expiry": [days_to_expiry],
        "Years to Expiry": [years_to_expiry]
    })

    calls = ticker.option_chain(expiry).calls
    if "lastTradeDate" in calls.columns:
        calls["lastTradeDate"] = pd.to_datetime(calls["lastTradeDate"]).dt.date
        cutoff_date = (TODAY - timedelta(days=30)).date()
        calls = calls[calls["lastTradeDate"] >= cutoff_date]

    if calls.empty:
        return summary_df, None, []

    cols = ["strike", "lastPrice", "bid", "ask", "volume", "openInterest", "impliedVolatility", "inTheMoney"]
    calls = calls[cols].copy()
    calls["middle"] = ((calls["bid"] + calls["ask"]) / 2).round(2)
    calls["impliedVolatility"] = (calls["impliedVolatility"]*100).round(2)

    percentages = [5, 10, 15, 20, 25, 30, 35, 40]
    column_names = []
    for pct in percentages:
        target = round(last_price * ((1 + pct/100) ** years_to_expiry), 2)
        col_name = f"{ticker_input}: {pct}% - {target}"
        column_names.append(col_name)
        calls[col_name] = ((target - (calls["strike"] + calls["middle"])) / calls["middle"]) * 100
        calls[col_name] = calls[col_name].round(2).astype(str) + "%"

    front_cols = ["strike", "middle", "volume"]
    other_cols = [c for c in calls.columns if c not in front_cols]
    calls = calls[front_cols + other_cols]

    return summary_df, calls, column_names

# --- Fetch data ---
if ticker_input and expiry and (refresh_pressed or st.session_state.calls_data is None):
    with st.spinner("Loading options data..."):
        summary_df, calls, column_names = fetch_option_data_cached(ticker_input, expiry)
    st.session_state.calls_data = calls
    st.session_state.summary_data = summary_df
    st.session_state.column_names = column_names
else:
    calls = st.session_state.calls_data
    summary_df = st.session_state.summary_data
    column_names = st.session_state.column_names or []

# --- Show summary ---
if summary_df is not None:
    st.subheader("Summary")
    st.dataframe(summary_df.round(2))

# --- Call Options section ---
if calls is not None and not calls.empty:
    st.subheader("Call Options")

    show_active_only = st.checkbox(
        "Show only active contracts (volume > 0 or OI > 10)",
        value=True,
        key="show_active_only"
    )

    # Apply active filter dynamically
    display_calls = calls.copy()
    if show_active_only:
        display_calls = display_calls[
            (display_calls["volume"] > 0) | (display_calls["openInterest"] > 10)
        ]

    if display_calls.empty:
        st.warning("No active contracts found with current filter.")
    else:
        # Column toggles
        col1, col2, col3, col4 = st.columns(4)
        col5, col6, col7, col8 = st.columns(4)
        checkbox_columns = [col1, col2, col3, col4, col5, col6, col7, col8]

        show_cols = []
        defaults = [True, True, True, True, False, False, False, False]

        for i, col_name in enumerate(column_names):
            show = checkbox_columns[i].checkbox(
                label=col_name,
                value=defaults[i],
                key=f"show_{col_name}"
            )
            if show:
                show_cols.append(col_name)

        # Visible columns for table
        display_cols = ["strike", "middle", "volume"] + show_cols + \
                       ["lastPrice", "bid", "ask", "openInterest", "impliedVolatility", "inTheMoney"]

        display_df = display_calls[display_cols].copy()

        # Build AgGrid
        gb = GridOptionsBuilder.from_dataframe(display_df)
        for col in ["strike", "middle", "volume"]:
            gb.configure_column(col, pinned='left', cellStyle={'fontWeight': 'bold'})

        for col in show_cols:
            if col in display_df.columns:
                col_numeric = display_df[col].str.rstrip('%').astype(float)
                max_val = col_numeric.replace([float('inf'), float('-inf')], pd.NA).max() or 0
                js_code = JsCode(f"""
                function(params) {{
                    if (!params.value) return {{}};
                    let val = parseFloat(params.value.replace('%',''));
                    if (isNaN(val)) return {{}};
                    if (val < 0) {{
                        return {{'color':'white','backgroundColor':'#5DADE2'}};
                    }} else if (val === {max_val}) {{
                        return {{'color':'black','backgroundColor':'#F39C12'}};
                    }} else {{
                        return {{}};
                    }}
                }}
                """)
                gb.configure_column(col, cellStyle=js_code)

        grid_options = gb.build()
        AgGrid(
            display_df,
            gridOptions=grid_options,
            fit_columns_on_grid_load=True,
            height=600,
            width='100%',
            allow_unsafe_jscode=True
        )

        # ── Return % Chart + vertical line at max sum (only where all lines exist) ──
        st.subheader("Potential Positive Return % by Strike Price")

        if show_cols and len(show_cols) > 0:
            fig = go.Figure()

            colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                      '#8c564b', '#e377c2', '#7f7f7f'][:len(show_cols)]

            # Calculate intersection of ALL positive masks
            all_positive_mask = None
            col_returns = {}
            
            for col_name in show_cols:
                numeric_col = display_calls[col_name].str.rstrip('%').astype(float)
                positive_mask = (numeric_col > 0) & numeric_col.notna() & np.isfinite(numeric_col)
                col_returns[col_name] = numeric_col
                
                if all_positive_mask is None:
                    all_positive_mask = positive_mask
                else:
                    all_positive_mask &= positive_mask

            # Find strike with highest sum among intersection points
            best_strike = None
            max_sum_value = 0
            
            if all_positive_mask.any():
                valid_strikes = display_calls.loc[all_positive_mask, "strike"]
                for strike in valid_strikes:
                    strike_idx = display_calls[display_calls["strike"] == strike].index[0]
                    current_sum = sum(col_returns[col_name][strike_idx] for col_name in show_cols)
                    if current_sum > max_sum_value:
                        max_sum_value = current_sum
                        best_strike = strike

            # Add vertical line showing only the strike
            if best_strike is not None and max_sum_value > 0:
                fig.add_vline(
                    x=best_strike,
                    line_width=3,
                    line_dash="dash",
                    line_color="#00FF00",
                    annotation_text=f"Strike: {best_strike}",
                    annotation_position="top",
                    annotation_font_size=14,
                    annotation_font_color="#00FF00"
                )

            # Plot individual lines
            for i, col_name in enumerate(show_cols):
                pct_label = col_name.split(": ")[1].split(" - ")[0]
                numeric_col = display_calls[col_name].str.rstrip('%').astype(float)
                positive_mask = (numeric_col > 0) & numeric_col.notna() & np.isfinite(numeric_col)

                if positive_mask.any():
                    strikes = display_calls.loc[positive_mask, "strike"]
                    returns = numeric_col[positive_mask]

                    fig.add_trace(go.Scatter(
                        x=strikes,
                        y=returns,
                        mode='lines+markers',
                        name=f"{pct_label}",
                        line=dict(color=colors[i % len(colors)], width=2),
                        marker=dict(size=6),
                        hovertemplate='Strike: %{x}<br>Return: %{y:.1f}%<extra></extra>'
                    ))

            fig.update_layout(
                xaxis_title="Strike Price",
                yaxis_title="Potential Return (%)",
                hovermode="x unified",
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                height=500,
                template="plotly_dark" if st.get_option("theme.base") == "dark" else "plotly_white",
                margin=dict(l=40, r=40, t=80, b=60)
            )

            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Select at least one percentage column above to show the return chart.")

    # --- Excel download ---
    if calls is not None:
        blue_fill = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
        orange_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="Calls", index=False, startrow=0)
            calls_excel = calls.copy()
            for col in column_names:
                if col in calls_excel.columns:
                    calls_excel[col] = calls_excel[col].str.rstrip('%').astype(float)
            calls_excel.to_excel(writer, sheet_name="Calls", index=False, startrow=6)
        output.seek(0)

        wb = load_workbook(output)
        ws = wb["Calls"]
        ws.freeze_panes = ws['D7']

        for cell in ws[7]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        header_values = [cell.value for cell in ws[7]]
        for col_name in column_names:
            if col_name not in header_values:
                continue
            col_idx = header_values.index(col_name) + 1
            ws.conditional_formatting.add(
                f"{get_column_letter(col_idx)}8:{get_column_letter(col_idx)}{ws.max_row}",
                CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=blue_fill)
            )
            col_cells = [ws.cell(row=row, column=col_idx).value for row in range(8, ws.max_row + 1)]
            numeric_vals = [v for v in col_cells if isinstance(v, (int, float))]
            if numeric_vals:
                max_val = max(numeric_vals)
                ws.conditional_formatting.add(
                    f"{get_column_letter(col_idx)}8:{get_column_letter(col_idx)}{ws.max_row}",
                    CellIsRule(operator='equal', formula=[str(max_val)], stopIfTrue=True, fill=orange_fill)
                )

        output2 = BytesIO()
        wb.save(output2)
        output2.seek(0)

        st.download_button(
            label="📥 Download Excel (formatted - all columns)",
            data=output2,
            file_name=f"{ticker_input}_CallOptions_{expiry}_formatted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ─── AI ANALYSIS SECTION ───────────────────────────────────────────
if calls is not None and summary_df is not None:
    st.markdown("---")
    st.header("🤖 AI Covered Call Analysis")
    
    # Initialize chat history in session state
    if "chat_history_calls" not in st.session_state:
        st.session_state.chat_history_calls = []
    if "initial_context_calls" not in st.session_state:
        st.session_state.initial_context_calls = None

    if not GEMINI_AVAILABLE:
        st.error("❌ Google Gemini package not installed")
        st.code("pip install google-generativeai", language="bash")
        st.info("Install the package and restart the app to enable AI analysis")
    elif not gemini_api_key:
        st.warning("⚠️ Enter your Gemini API key in the sidebar to enable AI analysis")
        st.info("Get a free API key at: https://aistudio.google.com/app/apikey")
    else:
        # Initial Analysis Button
        if st.button("🔍 Generate AI Covered Call Strategy", type="primary", use_container_width=True):
            with st.spinner("Analyzing options data with Gemini AI..."):
                try:
                    # Prepare context for AI
                    last_price = summary_df["Last Price"].iloc[0]
                    days_to_expiry = summary_df["Days to Expiry"].iloc[0]
                    
                    # Get top 5 contracts by volume
                    active_calls = calls[(calls["volume"] > 0) | (calls["openInterest"] > 10)].copy()
                    
                    # Convert percentage columns to numeric for analysis
                    for col in column_names[:4]:  # First 4 percentage columns
                        if col in active_calls.columns:
                            active_calls[f"{col}_numeric"] = active_calls[col].str.rstrip('%').astype(float)
                    
                    top_volume = active_calls.nlargest(5, 'volume')[
                        ['strike', 'middle', 'volume', 'openInterest', 'impliedVolatility']
                    ].to_string(index=False)
                    
                    # Find optimal strike from chart (if available)
                    optimal_strike_info = ""
                    if best_strike is not None:
                        optimal_strike_info = f"\nOptimal Strike (highest combined return): ${best_strike}"
                    
                    # Get IV percentile info
                    avg_iv = active_calls['impliedVolatility'].mean()
                    max_iv = active_calls['impliedVolatility'].max()
                    min_iv = active_calls['impliedVolatility'].min()
                    
                    context = f"""
Analyze this covered call opportunity and provide a clear recommendation.

TICKER: {ticker_input}
CURRENT STOCK PRICE: ${last_price:.2f}
EXPIRATION: {expiry} ({days_to_expiry} days)

IMPLIED VOLATILITY OVERVIEW:
- Average IV: {avg_iv:.1f}%
- Range: {min_iv:.1f}% to {max_iv:.1f}%

TOP 5 STRIKES BY VOLUME:
{top_volume}
{optimal_strike_info}

STRATEGY CONTEXT:
This is for selling covered calls - the investor owns the stock and wants to generate premium income while potentially selling shares at a profit.

Provide analysis in this format:

1. MARKET CONDITIONS: (2-3 sentences about current IV environment and what it means for premium sellers)

2. RECOMMENDED STRIKE(S):
   - Best conservative strike: [strike] - Premium: $[middle] - IV: [IV%]
   - Best aggressive strike: [strike] - Premium: $[middle] - IV: [IV%]
   Reasoning: (Why these strikes? Consider distance from current price, premium, and upside capture)

3. INCOME POTENTIAL:
   - Annualized return if called away: [calculate based on premium + capital gain]
   - Annualized return if not called: [based on premium only]

4. RISK ASSESSMENT:
   - Probability of assignment: [High/Medium/Low based on strike vs current price]
   - Upside capped at: [strike price, calculate % gain from current]
   - What to watch: [key price levels or events]

5. RECOMMENDATION: **SELL CALLS** or **WAIT**
   Reasoning: (2-3 sentences on timing, IV level, and strategy fit)

6. ACTION PLAN:
   - Suggested strike: $[price]
   - Order type: [limit/market]
   - Entry price: $[middle or slightly above bid]
   - Exit strategy: [when to roll or close]

Be specific with numbers. Focus on income generation and risk management for covered call sellers.
"""

                    # Store initial context for follow-ups
                    st.session_state.initial_context_calls = context
                    
                    # Call Gemini API
                    model = genai.GenerativeModel('gemini-2.5-flash')
                    response = model.generate_content(context)
                    
                    # Clear previous chat and add initial exchange
                    st.session_state.chat_history_calls = [
                        {"role": "user", "content": "Analyze these covered call options"},
                        {"role": "assistant", "content": response.text}
                    ]
                    
                    st.rerun()
                    
                except Exception as e:
                    st.error(f"Error generating analysis: {str(e)}")
                    st.info("Check your API key or try again. Error details above.")
        
        # Display chat history
        if st.session_state.chat_history_calls:
            st.markdown("### 💬 AI Conversation")
            
            # Display all messages
            for i, msg in enumerate(st.session_state.chat_history_calls):
                if msg["role"] == "user" and i > 0:  # Skip first generic prompt
                    with st.chat_message("user"):
                        st.markdown(msg["content"])
                elif msg["role"] == "assistant":
                    with st.chat_message("assistant"):
                        st.markdown(msg["content"])
            
            # Follow-up question input
            st.markdown("---")
            follow_up = st.text_input(
                "💭 Ask a follow-up question:",
                placeholder="e.g., What if IV drops by 20%? Should I roll to a higher strike?",
                key="follow_up_calls"
            )
            
            col1, col2 = st.columns([1, 5])
            with col1:
                send_button = st.button("Send", type="primary", use_container_width=True)
            with col2:
                if st.button("🗑️ Clear Conversation", use_container_width=True):
                    st.session_state.chat_history_calls = []
                    st.session_state.initial_context_calls = None
                    st.rerun()
            
            if send_button and follow_up:
                with st.spinner("Thinking..."):
                    try:
                        # Build conversation history for context
                        conversation = [{"role": "user", "parts": [st.session_state.initial_context_calls]}]
                        
                        for msg in st.session_state.chat_history_calls:
                            conversation.append({
                                "role": "user" if msg["role"] == "user" else "model",
                                "parts": [msg["content"]]
                            })
                        
                        # Add new question
                        conversation.append({"role": "user", "parts": [follow_up]})
                        
                        # Call Gemini with full conversation
                        model = genai.GenerativeModel('gemini-2.5-flash')
                        chat = model.start_chat(history=conversation[:-1])
                        response = chat.send_message(follow_up)
                        
                        # Add to chat history
                        st.session_state.chat_history_calls.append({"role": "user", "content": follow_up})
                        st.session_state.chat_history_calls.append({"role": "assistant", "content": response.text})
                        
                        st.rerun()
                        
                    except Exception as e:
                        st.error(f"Error: {str(e)}")
            
            # Disclaimer
            st.warning("⚠️ **Disclaimer**: This is AI-generated analysis for educational purposes only. Not financial advice. Options trading involves significant risk. Always do your own research and consult with a financial advisor.")