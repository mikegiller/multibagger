# options_explorer.py — shared Options Data Explorer logic for Stocks/ETFs and Indexes
#
# Fetch, grid, chart, and Excel export are identical and shared here. The two
# asset classes differ in three places:
#   - mid-price calc: index LEAPS chains routinely have zero-bid/zero-ask strikes
#     (wide, illiquid markets) that break a plain (bid+ask)/2 calc into "inf%"
#     return columns and a broken Excel export; indexes get a lastPrice fallback,
#     stocks/ETFs keep the plain calc since that failure mode is rare there.
#   - AI analysis framing: you can't own an index to sell covered calls against
#     it, so the index variant reframes the Gemini prompt as a long call/LEAPS
#     buying analysis instead.
#   - a VIX context line shown only for indexes.
import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import yfinance as yf
import utils  # applies the yfinance cache workaround on import
from io import BytesIO
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import plotly.graph_objects as go

TODAY = datetime.now()

ASSET_CONFIG = {
    "stocks": {
        "page_title": "Call Options Viewer",
        "heading": "📈 Call Options Viewer",
        "purpose": (
            "**Purpose:** Find the best-performing LEAPS and future options under different "
            "annual growth assumptions (5%, 10%, 15%, etc.) for **stocks and ETFs**. Useful "
            "for building long-term positions with defined risk."
        ),
        "favorites_category": "stocks",
        "use_lastprice_fallback": False,
        "ai_header": "🤖 AI Covered Call Analysis",
        "ai_button_label": "🔍 Generate AI Covered Call Strategy",
        "ai_price_label": "CURRENT STOCK PRICE",
        "ai_intro": "Analyze this covered call opportunity and provide a clear recommendation.",
        "ai_strategy_context": (
            "This is for selling covered calls - the investor owns the stock and wants to "
            "generate premium income while potentially selling shares at a profit."
        ),
        "ai_format_block": """1. MARKET CONDITIONS: (2-3 sentences about current IV environment and what it means for premium sellers)

2. RECOMMENDED STRIKE(S):
   - Best conservative strike: [strike] - Premium: $[middle] - IV: [IV%]
   - Best aggressive strike: [strike] - Premium: $[middle] - IV: [IV%]
   Reasoning: (Why these strikes? Consider distance from current price, premium, and upside capture)

3. INCOME POTENTIAL:
   - Annualized return if called away: [calculate based on premium + capital gain]
   - Annualized return if not called: [based on premium only]

4. RISK ASSESSMENT:
   - Probability of assignment: [High/Medium/Low based on strike vs current price]
   - Upside capped at: [strike price, calculate % gain from current]
   - What to watch: [key price levels or events]

5. RECOMMENDATION: **SELL CALLS** or **WAIT**
   Reasoning: (2-3 sentences on timing, IV level, and strategy fit)

6. ACTION PLAN:
   - Suggested strike: $[price]
   - Order type: [limit/market]
   - Entry price: $[middle or slightly above bid]
   - Exit strategy: [when to roll or close]""",
        "ai_focus_line": "Be specific with numbers. Focus on income generation and risk management for covered call sellers.",
        "ai_initial_user_msg": "Analyze these covered call options",
    },
    "index": {
        "page_title": "Index Options Viewer",
        "heading": "📈 Index Options Viewer",
        "purpose": (
            "**Purpose:** Find the best-performing LEAPS and future options under different "
            "annual growth assumptions (5%, 10%, 15%, etc.) for **indexes** (SPX, NDX, OEX, "
            "etc.). Useful for building long-term positions with defined risk. Illiquid "
            "far-dated LEAPS with no live bid/ask fall back to the last recent trade price "
            "instead of showing inf%."
        ),
        "favorites_category": "index",
        "use_lastprice_fallback": True,
        "ai_header": "🤖 AI Call-Buying Analysis",
        "ai_button_label": "🔍 Generate AI Call-Buying Strategy",
        "ai_price_label": "CURRENT INDEX LEVEL",
        "ai_intro": "Analyze this long call / LEAPS buying opportunity and provide a clear recommendation.",
        "ai_strategy_context": (
            "This is for buying long calls (or LEAPS) on a broad-market index as leveraged, "
            "capital-efficient exposure. The investor cannot own the index directly and is not "
            "selling covered calls against it — the goal is identifying the call strike/expiration "
            "with the best projected return if the index reaches the target growth level."
        ),
        "ai_format_block": """1. MARKET CONDITIONS: (2-3 sentences about current IV environment and what it means for call buyers)

2. RECOMMENDED STRIKE(S):
   - Best conservative (higher delta, less leverage) strike: [strike] - Premium: $[middle] - IV: [IV%]
   - Best aggressive (lower delta, more leverage) strike: [strike] - Premium: $[middle] - IV: [IV%]
   Reasoning: (Why these strikes? Consider delta, leverage, and cost relative to the projected move)

3. RETURN POTENTIAL:
   - Projected return if the index reaches the target growth level: [use optimal strike info]
   - Breakeven index level: [strike + premium]

4. RISK ASSESSMENT:
   - Max loss: 100% of premium paid if the index finishes below the strike at expiration
   - Time decay (theta) exposure: [High/Medium/Low based on days to expiry]
   - What to watch: [key index levels, VIX moves, or macro events]

5. RECOMMENDATION: **BUY CALLS** or **WAIT**
   Reasoning: (2-3 sentences on timing, IV level, and strategy fit)

6. ACTION PLAN:
   - Suggested strike: $[price]
   - Order type: [limit/market]
   - Entry price: $[middle or slightly below ask]
   - Exit strategy: [profit target or roll trigger]""",
        "ai_focus_line": "Be specific with numbers. Focus on capital efficiency and risk management for long call/LEAPS buyers.",
        "ai_initial_user_msg": "Analyze this LEAPS call-buying opportunity",
    },
}


@st.cache_data(ttl=600, show_spinner="Fetching fresh options data from Yahoo Finance...")
def fetch_option_data_cached(ticker_input, expiry, use_lastprice_fallback):
    ticker = yf.Ticker(ticker_input)
    hist = ticker.history(period="1d")
    last_price = round(hist["Close"].iloc[-1], 2) if not hist.empty else None
    expiry_date = datetime.strptime(expiry, "%Y-%m-%d")
    days_to_expiry = (expiry_date - TODAY).days
    years_to_expiry = days_to_expiry / 365 if days_to_expiry > 0 else 0.0001

    summary_df = pd.DataFrame({
        "Ticker": [ticker_input],
        "Last Price": [last_price],
        "Expiration": [expiry],
        "Days to Expiry": [days_to_expiry],
        "Years to Expiry": [years_to_expiry]
    })

    calls = ticker.option_chain(expiry).calls
    if "lastTradeDate" in calls.columns:
        calls["lastTradeDate"] = pd.to_datetime(calls["lastTradeDate"]).dt.date
        cutoff_date = (TODAY - timedelta(days=30)).date()
        calls = calls[calls["lastTradeDate"] >= cutoff_date]

    if calls.empty:
        return summary_df, None, [], []

    cols = ["strike", "lastPrice", "bid", "ask", "volume", "openInterest", "impliedVolatility", "inTheMoney"]
    if use_lastprice_fallback:
        cols = cols + ["lastTradeDate"]
    calls = calls[cols].copy()
    calls["middle"] = ((calls["bid"] + calls["ask"]) / 2).round(2)
    calls["impliedVolatility"] = (calls["impliedVolatility"]*100).round(2)

    if use_lastprice_fallback:
        # No bid/ask market (common on illiquid far-dated index LEAPS): fall back to
        # lastPrice if it's from a recent trade, otherwise leave middle at 0 so the
        # return% is N/A.
        recent_cutoff = (pd.Timestamp(TODAY.date()) - pd.tseries.offsets.BDay(3)).date()
        use_last_price = (calls["middle"] == 0) & (calls["lastPrice"] > 0) & (calls["lastTradeDate"] >= recent_cutoff)
        calls.loc[use_last_price, "middle"] = calls.loc[use_last_price, "lastPrice"]
        calls = calls.drop(columns=["lastTradeDate"])

    percentages = [5, 10, 15, 20, 25, 30, 35, 40]
    column_names = []
    actual_percentages = []
    has_middle = calls["middle"] > 0
    for pct in percentages:
        if years_to_expiry <= 1:
            actual_pct = pct
            target = round(last_price * (1 + pct / 100), 2)
        else:
            actual_pct = ((1 + pct / 100) ** years_to_expiry - 1) * 100
            target = round(last_price * ((1 + pct / 100) ** years_to_expiry), 2)
        actual_percentages.append(actual_pct)
        col_name = f"{ticker_input}: {pct}% - {target}"
        column_names.append(col_name)
        if use_lastprice_fallback:
            ret = pd.Series(np.nan, index=calls.index)
            ret[has_middle] = ((target - (calls.loc[has_middle, "strike"] + calls.loc[has_middle, "middle"])) / calls.loc[has_middle, "middle"]) * 100
            calls[col_name] = ret.round(2).astype(str) + "%"
            calls.loc[~has_middle, col_name] = "N/A"
        else:
            calls[col_name] = ((target - (calls["strike"] + calls["middle"])) / calls["middle"]) * 100
            calls[col_name] = calls[col_name].round(2).astype(str) + "%"

    front_cols = ["strike", "middle", "volume"]
    other_cols = [c for c in calls.columns if c not in front_cols]
    calls = calls[front_cols + other_cols]

    return summary_df, calls, column_names, actual_percentages


def render(asset_class):
    """Render the full Options Data Explorer page for asset_class ("stocks" or "index")."""
    cfg = ASSET_CONFIG[asset_class]
    use_lastprice_fallback = cfg["use_lastprice_fallback"]
    prefix = f"ode_{asset_class}_"

    def k(name):
        return prefix + name

    # --- Page settings ---
    st.set_page_config(page_title=cfg["page_title"], layout="wide")
    st.title(cfg["heading"])
    st.info(cfg["purpose"])
    if asset_class == "index":
        utils.vix_badge()
    st.write("Enter a ticker to view call options, toggle columns, and see return potential graph (positive only).")

    # --- Gemini API Key Setup (in sidebar) ──────────────────────────────
    gemini_client = utils.gemini_sidebar()

    # --- Session state ---
    _STATE_KEYS = [k("expiry"), k("calls_data"), k("summary_data"), k("column_names"),
                   k("actual_percentages"), k("last_fetched_ticker"), k("last_fetched_expiry")]
    for key in _STATE_KEYS:
        if key not in st.session_state:
            st.session_state[key] = None

    # --- Reset button ---
    if st.button("🔄 Reset", key=k("reset_btn")):
        for key in _STATE_KEYS:
            st.session_state[key] = None

    # --- Favorites ---
    _favs = utils.load_favorites(cfg["favorites_category"])

    if "fav" in st.query_params and st.query_params["fav"] in _favs:
        st.session_state[k("ticker")] = st.query_params["fav"]
        utils.set_active_ticker(st.query_params["fav"], category=cfg["favorites_category"])
        st.session_state[k("calls_data")] = None
        st.session_state[k("expiry")] = None
        del st.query_params["fav"]

    st.markdown(utils.favorites_html(_favs), unsafe_allow_html=True)

    # --- Ticker input ---
    ticker_input = utils.ticker_input(
        k("ticker"), category=cfg["favorites_category"], label="Enter ticker symbol (required):"
    )
    utils.warn_if_wrong_report(
        ticker_input, asset_class,
        stocks_page="pages/Options_Data_Explorer_Stocks.py",
        index_page="pages/Options_Data_Explorer_Index.py",
    )
    # Clear cached data when ticker changes
    if ticker_input != st.session_state.get(k("_prev_ticker")):
        st.session_state[k("calls_data")] = None
        st.session_state[k("expiry")] = None
    st.session_state[k("_prev_ticker")] = ticker_input

    # --- Fetch expirations ---
    expirations = []
    if ticker_input:
        try:
            ticker = yf.Ticker(ticker_input)
            hist = ticker.history(period="1d")
            if hist.empty:
                st.error("No data found for ticker.")
            elif not ticker.options:
                st.error("No options available for this ticker.")
            else:
                expirations = ticker.options
        except Exception as e:
            st.error(f"Error fetching ticker info: {e}")

    # --- Expiration dropdown with days to expiry ---
    expiry = None
    if expirations:
        # Sort descending (furthest expiration first)
        expirations_sorted = sorted(expirations, reverse=True)

        # Format options with days remaining
        expiry_options = ["Select expiration date..."]
        for exp_date_str in expirations_sorted:
            exp_date = datetime.strptime(exp_date_str, "%Y-%m-%d")
            days_left = (exp_date - TODAY).days
            label = f"{exp_date_str} ({days_left} days to expiration)"
            expiry_options.append(label)

        # Find index of previously selected expiry
        selected_index = 0
        if st.session_state[k("expiry")]:
            prev_label = f"{st.session_state[k('expiry')]} ({(datetime.strptime(st.session_state[k('expiry')], '%Y-%m-%d') - TODAY).days} days to expiration)"
            if prev_label in expiry_options:
                selected_index = expiry_options.index(prev_label)

        selected_label = st.selectbox(
            "Select expiration date (furthest first):",
            expiry_options,
            index=selected_index,
            key=k("expiry_select"),
        )

        # Extract the date part back from selected label
        if selected_label != "Select expiration date...":
            expiry = selected_label.split(" ")[0]  # e.g. "2026-01-16"
        else:
            expiry = None

        st.session_state[k("expiry")] = expiry

    # --- Data Refresh button ---
    refresh_pressed = st.button("🔄 Data Refresh", key=k("refresh_btn"))

    # Clear cached data when expiry changes
    if expiry != st.session_state.get(k("last_fetched_expiry")) and expiry is not None:
        st.session_state[k("calls_data")] = None

    # --- Fetch data ---
    if ticker_input and expiry and (refresh_pressed or st.session_state[k("calls_data")] is None or not st.session_state[k("actual_percentages")]):
        with st.spinner("Loading options data..."):
            summary_df, calls, column_names, actual_percentages = fetch_option_data_cached(
                ticker_input, expiry, use_lastprice_fallback
            )
        st.session_state[k("calls_data")] = calls
        st.session_state[k("summary_data")] = summary_df
        st.session_state[k("column_names")] = column_names
        st.session_state[k("actual_percentages")] = actual_percentages
        st.session_state[k("last_fetched_ticker")] = ticker_input
        st.session_state[k("last_fetched_expiry")] = expiry
    else:
        calls = st.session_state[k("calls_data")]
        summary_df = st.session_state[k("summary_data")]
        column_names = st.session_state[k("column_names")] or []
        actual_percentages = st.session_state[k("actual_percentages")] or []

    # --- Show summary ---
    if summary_df is not None:
        st.subheader("Summary")
        st.dataframe(summary_df.round(2))

    # Optimal strike from the return chart; stays None if the chart didn't render
    best_strike = None

    # --- Call Options section ---
    if calls is not None and not calls.empty:
        st.subheader("Call Options")

        show_active_only = st.checkbox(
            "Show only active contracts (volume > 0 or OI > 10)",
            value=True,
            key=k("show_active_only")
        )

        # Apply active filter dynamically
        display_calls = calls.copy()
        if show_active_only:
            display_calls = display_calls[
                (display_calls["volume"] > 0) | (display_calls["openInterest"] > 10)
            ]

        if display_calls.empty:
            st.warning("No active contracts found with current filter.")
        else:
            # Column toggles
            col1, col2, col3, col4 = st.columns(4)
            col5, col6, col7, col8 = st.columns(4)
            checkbox_columns = [col1, col2, col3, col4, col5, col6, col7, col8]

            show_cols = []
            defaults = [True, True, False, False, False, False, False, False]
            days_to_expiry = int(summary_df["Days to Expiry"].iloc[0]) if summary_df is not None else 0

            for i, col_name in enumerate(column_names):
                show = checkbox_columns[i].checkbox(
                    label=col_name,
                    value=defaults[i],
                    key=k(f"show_{col_name}")
                )
                if actual_percentages and i < len(actual_percentages):
                    checkbox_columns[i].caption(f"= {actual_percentages[i]:.2f}% total")
                if show:
                    show_cols.append(col_name)

            # Visible columns for table
            display_cols = ["strike", "middle", "volume"] + show_cols + \
                           ["lastPrice", "bid", "ask", "openInterest", "impliedVolatility", "inTheMoney"]

            display_df = display_calls[display_cols].copy()

            # Build AgGrid
            gb = GridOptionsBuilder.from_dataframe(display_df)
            for col in ["strike", "middle", "volume"]:
                gb.configure_column(col, pinned='left', cellStyle={'fontWeight': 'bold'})

            for col in show_cols:
                if col in display_df.columns:
                    col_numeric = pd.to_numeric(display_df[col].str.rstrip('%'), errors='coerce')
                    max_val = col_numeric.replace([float('inf'), float('-inf')], pd.NA).max()
                    if pd.isna(max_val):
                        max_val = 0
                    js_code = JsCode(f"""
                    function(params) {{
                        if (!params.value) return {{}};
                        let val = parseFloat(params.value.replace('%',''));
                        if (isNaN(val)) return {{}};
                        if (val < 0) {{
                            return {{'color':'white','backgroundColor':'#5DADE2'}};
                        }} else if (val === {max_val}) {{
                            return {{'color':'black','backgroundColor':'#F39C12'}};
                        }} else {{
                            return {{}};
                        }}
                    }}
                    """)
                    gb.configure_column(col, cellStyle=js_code)

            grid_options = gb.build()
            AgGrid(
                display_df,
                gridOptions=grid_options,
                fit_columns_on_grid_load=True,
                height=600,
                width='100%',
                allow_unsafe_jscode=True
            )

            # ── Return % Chart + vertical line at max sum (only where all lines exist) ──
            st.subheader("Potential Positive Return % by Strike Price")

            if show_cols and len(show_cols) > 0:
                fig = go.Figure()

                colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                          '#8c564b', '#e377c2', '#7f7f7f'][:len(show_cols)]

                # Calculate intersection of ALL positive masks
                all_positive_mask = None
                col_returns = {}

                for col_name in show_cols:
                    numeric_col = pd.to_numeric(display_calls[col_name].str.rstrip('%'), errors='coerce')
                    positive_mask = (numeric_col > 0) & numeric_col.notna() & np.isfinite(numeric_col)
                    col_returns[col_name] = numeric_col

                    if all_positive_mask is None:
                        all_positive_mask = positive_mask
                    else:
                        all_positive_mask &= positive_mask

                # Find strike with highest sum among intersection points
                best_strike = None
                max_sum_value = 0

                if all_positive_mask.any():
                    valid_strikes = display_calls.loc[all_positive_mask, "strike"]
                    for strike in valid_strikes:
                        strike_idx = display_calls[display_calls["strike"] == strike].index[0]
                        current_sum = sum(col_returns[col_name][strike_idx] for col_name in show_cols)
                        if current_sum > max_sum_value:
                            max_sum_value = current_sum
                            best_strike = strike

                # Add vertical line showing only the strike
                if best_strike is not None and max_sum_value > 0:
                    fig.add_vline(
                        x=best_strike,
                        line_width=3,
                        line_dash="dash",
                        line_color="#00FF00",
                        annotation_text=f"Strike: {best_strike}",
                        annotation_position="top",
                        annotation_font_size=14,
                        annotation_font_color="#00FF00"
                    )

                # Plot individual lines
                for i, col_name in enumerate(show_cols):
                    pct_label = col_name.split(": ")[1].split(" - ")[0]
                    numeric_col = pd.to_numeric(display_calls[col_name].str.rstrip('%'), errors='coerce')
                    positive_mask = (numeric_col > 0) & numeric_col.notna() & np.isfinite(numeric_col)

                    if positive_mask.any():
                        strikes = display_calls.loc[positive_mask, "strike"]
                        returns = numeric_col[positive_mask]

                        fig.add_trace(go.Scatter(
                            x=strikes,
                            y=returns,
                            mode='lines+markers',
                            name=f"{pct_label}",
                            line=dict(color=colors[i % len(colors)], width=2),
                            marker=dict(size=6),
                            hovertemplate='Strike: %{x}<br>Return: %{y:.1f}%<extra></extra>'
                        ))

                fig.update_layout(
                    xaxis_title="Strike Price",
                    yaxis_title="Potential Return (%)",
                    hovermode="x unified",
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                    height=500,
                    template="plotly_dark" if st.get_option("theme.base") == "dark" else "plotly_white",
                    margin=dict(l=40, r=40, t=80, b=60)
                )

                st.plotly_chart(fig, width='stretch')
            else:
                st.info("Select at least one percentage column above to show the return chart.")

        # --- Excel download ---
        if calls is not None:
            blue_fill = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
            orange_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")

            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                summary_df.to_excel(writer, sheet_name="Calls", index=False, startrow=0)
                calls_excel = calls.copy()
                for col in column_names:
                    if col in calls_excel.columns:
                        calls_excel[col] = pd.to_numeric(calls_excel[col].str.rstrip('%'), errors='coerce')
                calls_excel.to_excel(writer, sheet_name="Calls", index=False, startrow=6)
            output.seek(0)

            wb = load_workbook(output)
            ws = wb["Calls"]
            ws.freeze_panes = ws['D7']

            for cell in ws[7]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

            header_values = [cell.value for cell in ws[7]]
            for col_name in column_names:
                if col_name not in header_values:
                    continue
                col_idx = header_values.index(col_name) + 1
                ws.conditional_formatting.add(
                    f"{get_column_letter(col_idx)}8:{get_column_letter(col_idx)}{ws.max_row}",
                    CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=blue_fill)
                )
                col_cells = [ws.cell(row=row, column=col_idx).value for row in range(8, ws.max_row + 1)]
                numeric_vals = [v for v in col_cells if isinstance(v, (int, float))]
                if numeric_vals:
                    max_val = max(numeric_vals)
                    ws.conditional_formatting.add(
                        f"{get_column_letter(col_idx)}8:{get_column_letter(col_idx)}{ws.max_row}",
                        CellIsRule(operator='equal', formula=[str(max_val)], stopIfTrue=True, fill=orange_fill)
                    )

            output2 = BytesIO()
            wb.save(output2)
            output2.seek(0)

            st.download_button(
                label="📥 Download Excel (formatted - all columns)",
                data=output2,
                file_name=f"{ticker_input}_CallOptions_{expiry}_formatted.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=k("download_excel"),
            )

    # ─── AI ANALYSIS SECTION ───────────────────────────────────────────
    if calls is not None and summary_df is not None:
        st.markdown("---")
        st.header(cfg["ai_header"])

        # Initialize chat history in session state
        if k("chat_history_calls") not in st.session_state:
            st.session_state[k("chat_history_calls")] = []
        if k("initial_context_calls") not in st.session_state:
            st.session_state[k("initial_context_calls")] = None

        if not utils.GEMINI_AVAILABLE:
            st.error("❌ Google Gemini package not installed")
            st.code("pip install google-genai", language="bash")
            st.info("Install the package and restart the app to enable AI analysis")
        elif gemini_client is None:
            st.warning("⚠️ Enter your Gemini API key in the sidebar to enable AI analysis")
            st.info("Get a free API key at: https://aistudio.google.com/app/apikey")
        else:
            # Initial Analysis Button
            if st.button(cfg["ai_button_label"], type="primary", width='stretch', key=k("gen_ai_btn")):
                with st.spinner("Analyzing options data with Gemini AI..."):
                    try:
                        # Prepare context for AI
                        last_price = summary_df["Last Price"].iloc[0]
                        days_to_expiry = summary_df["Days to Expiry"].iloc[0]

                        # Get top 5 contracts by volume
                        active_calls = calls[(calls["volume"] > 0) | (calls["openInterest"] > 10)].copy()

                        # Convert percentage columns to numeric for analysis
                        for col in column_names[:4]:  # First 4 percentage columns
                            if col in active_calls.columns:
                                active_calls[f"{col}_numeric"] = pd.to_numeric(active_calls[col].str.rstrip('%'), errors='coerce')

                        top_volume = active_calls.nlargest(5, 'volume')[
                            ['strike', 'middle', 'volume', 'openInterest', 'impliedVolatility']
                        ].to_string(index=False)

                        # Find optimal strike from chart (if available)
                        optimal_strike_info = ""
                        if best_strike is not None:
                            optimal_strike_info = f"\nOptimal Strike (highest combined return): ${best_strike}"

                        # Get IV percentile info
                        avg_iv = active_calls['impliedVolatility'].mean()
                        max_iv = active_calls['impliedVolatility'].max()
                        min_iv = active_calls['impliedVolatility'].min()

                        context = f"""
{cfg["ai_intro"]}

TICKER: {ticker_input}
{cfg["ai_price_label"]}: ${last_price:.2f}
EXPIRATION: {expiry} ({days_to_expiry} days)

IMPLIED VOLATILITY OVERVIEW:
- Average IV: {avg_iv:.1f}%
- Range: {min_iv:.1f}% to {max_iv:.1f}%

TOP 5 STRIKES BY VOLUME:
{top_volume}
{optimal_strike_info}

STRATEGY CONTEXT:
{cfg["ai_strategy_context"]}

Provide analysis in this format:

{cfg["ai_format_block"]}

{cfg["ai_focus_line"]}
"""

                        # Store initial context for follow-ups
                        st.session_state[k("initial_context_calls")] = context

                        # Call Gemini API
                        response_text = utils.gemini_generate(gemini_client, context)

                        # Clear previous chat and add initial exchange
                        st.session_state[k("chat_history_calls")] = [
                            {"role": "user", "content": cfg["ai_initial_user_msg"]},
                            {"role": "assistant", "content": response_text}
                        ]

                        st.rerun()

                    except Exception as e:
                        st.error(f"Error generating analysis: {str(e)}")
                        st.info("Check your API key or try again. Error details above.")

            # Display chat history
            if st.session_state[k("chat_history_calls")]:
                st.markdown("### 💬 AI Conversation")

                # Display all messages
                for i, msg in enumerate(st.session_state[k("chat_history_calls")]):
                    if msg["role"] == "user" and i > 0:  # Skip first generic prompt
                        with st.chat_message("user"):
                            st.markdown(msg["content"])
                    elif msg["role"] == "assistant":
                        with st.chat_message("assistant"):
                            st.markdown(msg["content"])

                # Follow-up question input
                st.markdown("---")
                follow_up = st.text_input(
                    "💭 Ask a follow-up question:",
                    placeholder="e.g., What if IV drops by 20%? Should I roll to a higher strike?",
                    key=k("follow_up_calls")
                )

                col1, col2 = st.columns([1, 5])
                with col1:
                    send_button = st.button("Send", type="primary", width='stretch', key=k("send_btn"))
                with col2:
                    if st.button("🗑️ Clear Conversation", width='stretch', key=k("clear_btn")):
                        st.session_state[k("chat_history_calls")] = []
                        st.session_state[k("initial_context_calls")] = None
                        st.rerun()

                if send_button and follow_up:
                    with st.spinner("Thinking..."):
                        try:
                            reply = utils.gemini_chat_reply(
                                gemini_client,
                                st.session_state[k("initial_context_calls")],
                                st.session_state[k("chat_history_calls")],
                                follow_up,
                            )

                            # Add to chat history
                            st.session_state[k("chat_history_calls")].append({"role": "user", "content": follow_up})
                            st.session_state[k("chat_history_calls")].append({"role": "assistant", "content": reply})

                            st.rerun()

                        except Exception as e:
                            st.error(f"Error: {str(e)}")

                # Disclaimer
                st.warning("⚠️ **Disclaimer**: This is AI-generated analysis for educational purposes only. Not financial advice. Options trading involves significant risk. Always do your own research and consult with a financial advisor.")
