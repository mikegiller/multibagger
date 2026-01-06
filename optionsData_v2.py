import yfinance as yf
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import sys

try:
    from pick import pick
    use_pick = True
except ImportError:
    use_pick = False

# --- Function to safely get ticker ---
def get_ticker():
    while True:
        ticker_input = input("Enter ticker symbol (default = SPY): ").strip().upper() or "SPY"
        try:
            ticker = yf.Ticker(ticker_input)
            hist = ticker.history(period="1d")
            if hist.empty:
                raise ValueError("No data found for ticker.")
            if not ticker.options:
                raise ValueError("No options available for this ticker.")
            return ticker_input, ticker
        except Exception as e:
            print(f"❌ Invalid ticker '{ticker_input}'. {str(e)} Try again.\n")

# --- Get valid ticker ---
ticker_input, ticker = get_ticker()

# --- Last price ---
last_price = round(ticker.history(period="1d")["Close"].iloc[-1], 2)

# --- Expiration pick list ---
expirations = ticker.options
if not expirations:
    print("❌ No option expirations found for this ticker.")
    sys.exit(1)

print("\nAvailable expiration dates:")
if use_pick:
    title = "Select expiration date:"
    expiry, _ = pick(expirations, title)
else:
    for i, exp in enumerate(expirations, start=1):
        print(f"{i}. {exp}")
    while True:
        try:
            choice = int(input(f"\nEnter number of expiry (1-{len(expirations)}): "))
            if 1 <= choice <= len(expirations):
                expiry = expirations[choice - 1]
                break
            else:
                print("❌ Invalid number. Try again.")
        except ValueError:
            print("❌ Please enter a valid number.")

# --- Calculate expiration details ---
expiry_date = datetime.strptime(expiry, "%Y-%m-%d")
today = datetime.today()
days_to_expiry = (expiry_date - today).days
years_to_expiry = round(days_to_expiry / 365, 2)
years_to_expiry_orig = years_to_expiry

# --- Summary ---
summary_df = pd.DataFrame({
    "Ticker": [ticker_input],
    "Last Price": [last_price],
    "Expiration": [expiry],
    "Days to Expiry": [days_to_expiry],
    "Years to Expiry": [years_to_expiry]
})
print(f"\nSummary:\n{summary_df.to_string(index=False)}")

# --- Adjust for short-term options (after summary) ---
if years_to_expiry < 1:
    years_to_expiry = 1

# --- Fetch Calls ---
try:
    calls = ticker.option_chain(expiry).calls
except Exception as e:
    print(f"❌ Could not retrieve call options for {expiry}: {e}")
    sys.exit(1)

# --- Filter calls: only keep rows with lastTradeDate in last 30 days ---
cutoff_date = today - timedelta(days=30)
if "lastTradeDate" in calls.columns:
    # ensure timezone removed for comparison
    calls["lastTradeDate"] = calls["lastTradeDate"].dt.tz_localize(None)
    calls = calls[calls["lastTradeDate"] >= cutoff_date]

if calls.empty:
    print("⚠️ No call options traded in the last 30 days for this expiration.")
    sys.exit(0)  # exit gracefully

# --- Keep relevant columns ---
cols = ["strike","lastPrice","bid","ask","volume","openInterest","impliedVolatility","inTheMoney"]
calls = calls[cols].copy()
calls["middle"] = (calls["bid"] + calls["ask"]) / 2
calls["middle"] = calls["middle"].round(2)
calls["impliedVolatility"] = (calls["impliedVolatility"]*100).round(2)

cols_with_middle = ["strike","lastPrice","bid","ask","middle","volume","openInterest","impliedVolatility","inTheMoney"]
calls = calls[cols_with_middle]

# --- Dynamic % columns (with price target in header) ---
percentages = [5, 10, 15, 20, 25, 30, 35, 40]
column_names = []  # store full column names for later formatting use

for pct in percentages:
    percent_target = round(last_price * ((1 + pct / 100) ** years_to_expiry), 2)
    col_name = f"{ticker_input}: {pct}% - {percent_target}"
    column_names.append(col_name)

    calls[col_name] = ((percent_target - (calls["strike"] + calls["middle"])) / calls["middle"])
    calls[col_name] = calls[col_name].where(calls["middle"] != 0)
    calls[col_name] = calls[col_name].round(4)

# --- Save Excel ---
filename = f"{ticker_input}_CallOptions_{expiry}.xlsx"
with pd.ExcelWriter(filename, engine="openpyxl") as writer:
    summary_df.to_excel(writer, sheet_name="Calls", index=False, startrow=0)
    calls.to_excel(writer, sheet_name="Calls", index=False, startrow=6)

# --- Format Excel ---
wb = load_workbook(filename)
ws = wb["Calls"]

# Header formatting
for cell in ws[7]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
ws.freeze_panes = ws['A8']

# Autofit columns
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value is not None:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

# Conditional formatting
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

for col_name in column_names:
    col_idx = None
    for i, cell in enumerate(ws[7], start=1):
        if cell.value == col_name:
            col_idx = i
            break
    if col_idx is None:
        continue

    col_cells = [cell for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=8, max_row=ws.max_row) for cell in col]
    for cell in col_cells:
        if cell.value is not None:
            cell.number_format = "0.0%"

    ws.conditional_formatting.add(
        f"{ws.cell(row=8, column=col_idx).coordinate}:{ws.cell(row=ws.max_row, column=col_idx).coordinate}",
        CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill)
    )

    positive_values = [cell.value for cell in col_cells if cell.value not in (None, '') and cell.value > 0]
    if positive_values:
        max_val = max(positive_values)
        ws.conditional_formatting.add(
            f"{ws.cell(row=8, column=col_idx).coordinate}:{ws.cell(row=ws.max_row, column=col_idx).coordinate}",
            CellIsRule(operator='equal', formula=[str(max_val)], stopIfTrue=True, fill=green_fill)
        )

wb.save(filename)
print(f"✅ Excel file created: {filename}")

# --- PDF Report ---
pdf_file = f"{ticker_input}_CallOptions_{expiry}.pdf"
with PdfPages(pdf_file) as pdf:
    fig, ax = plt.subplots(figsize=(12, max(6, 0.3*len(calls)+4)))
    ax.axis('off')

    summary_lines = [
        f"Ticker: {ticker_input}",
        f"Last Price: {last_price}",
        f"Expiration: {expiry}",
        f"Days to Expiry: {days_to_expiry}",
        f"Years to Expiry: {years_to_expiry_orig:.4f}"
    ]
    plt.text(0, 1.02, "\n".join(summary_lines), fontsize=10, fontweight='bold', va='top', ha='left', transform=ax.transAxes)

    # prepare a df for PDF display
    df_pdf = calls.copy()

    # Format percentage-style columns using the actual column names (with price targets)
    for col_name in column_names:
        if col_name in df_pdf.columns:
            df_pdf[col_name] = df_pdf[col_name].apply(lambda x: f"{x*100:.1f}%" if pd.notnull(x) else "")

    # Build cell colors: iterate df_pdf.columns so columns remain in same order for table
    cell_colors = []
    for col in df_pdf.columns:
        col_colors = []
        if col in column_names:
            # use calls dataframe (which contains numeric values) to determine colors
            col_values = calls[col].tolist() if col in calls.columns else [None]*len(df_pdf)
            max_val = max([v for v in col_values if v not in (None, '') and v > 0], default=None)
            for v in col_values:
                if v is None:
                    col_colors.append('#FFFFFF')
                elif v < 0:
                    col_colors.append('#FFC7CE')
                elif (max_val is not None) and (v == max_val):
                    col_colors.append('#C6EFCE')
                else:
                    col_colors.append('#FFFFFF')
        else:
            # non-percent columns: keep white background
            col_colors = ['#FFFFFF']*len(df_pdf)
        cell_colors.append(col_colors)
    # transpose to match table API (rows x cols)
    cell_colors_t = list(map(list, zip(*cell_colors))) if cell_colors else [['#FFFFFF']*len(df_pdf.columns) for _ in range(len(df_pdf))]

    # Draw table
    table_data = df_pdf.values
    tbl = ax.table(cellText=table_data,
                   colLabels=df_pdf.columns,
                   cellColours=cell_colors_t,
                   loc='center',
                   cellLoc='center')
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(8)
    tbl.scale(1, 1.2)

    plt.title(f"{ticker_input} CALL Options - Expiry {expiry}", fontsize=12, fontweight='bold')
    pdf.savefig(bbox_inches='tight')
    plt.close()

print(f"✅ PDF report created: {pdf_file}")
