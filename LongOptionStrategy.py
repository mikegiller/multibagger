#!/usr/bin/env python3
"""
long_option_analysis.py

Fetches options chain for a ticker and evaluates expected profit for
long call or put options based on a user-specified expected stock move.
Exports results to a formatted Excel file.
"""

import sys
import datetime as dt
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import yfinance as yf
except ImportError:
    print("Please install yfinance: pip install yfinance")
    sys.exit(1)

# --- Option math helpers ---
def norm_cdf(x):
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))

def black_scholes_d1_d2(S, K, r, sigma, T):
    if sigma <= 0 or T <= 0:
        return None, None
    d1 = (math.log(S / K) + (r + 0.5 * sigma * sigma) * T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)
    return d1, d2

def option_mid_price(row):
    bid = row.get("bid")
    ask = row.get("ask")
    last = row.get("lastPrice")
    if bid is None or ask is None or (bid == 0 and ask == 0):
        return last if last is not None else 0.0
    return (bid + ask) / 2.0

# --- Build long option DataFrame ---
def build_long_options(option_type, chain_df, S0, expiry_date, today, expected_move_pct=0.05, r=0.0):
    T_days = (expiry_date - today).days
    if T_days <= 0:
        return pd.DataFrame()
    T = T_days / 365.0
    rows = []

    for _, row in chain_df.iterrows():
        strike = row['strike']
        mid_price = option_mid_price(row)
        debit = mid_price
        # Expected stock price
        ST_expected = S0 * (1 + expected_move_pct) if option_type == 'calls' else S0 * (1 - expected_move_pct)
        if option_type == 'calls':
            profit = max(ST_expected - strike, 0) - debit
        else:
            profit = max(strike - ST_expected, 0) - debit
        return_ratio = profit / debit if debit != 0 else float('inf')

        # Approx probability ITM using Black-Scholes
        iv = row.get('impliedVol') or 0.0
        try:
            d1, d2 = black_scholes_d1_d2(S0, strike, r, iv, T)
            prob_ITM = 1 - norm_cdf(d2) if option_type == 'calls' else norm_cdf(-d2) if d2 is not None else None
        except:
            prob_ITM = None

        rows.append({
            'option_type': option_type,
            'strike': strike,
            'mid_price': round(mid_price,2),
            'debit': round(debit,2),
            'expected_profit': round(profit,2),
            'return_ratio': round(return_ratio,4),
            'prob_ITM': round(prob_ITM,4) if prob_ITM is not None else None,
            'expiry': expiry_date.strftime("%Y-%m-%d"),
            'T_days': T_days,
            'iv': round(iv,4)
        })
    return pd.DataFrame(rows)

# --- Excel export ---
def export_to_excel(df, output_file):
    # Remove timezone info
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = pd.to_datetime(df[col]).dt.tz_localize(None)

    wb = Workbook()
    ws = wb.active
    ws.title = "long_options"

    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    pct_style = NamedStyle(name="percent_style", number_format="0.0%")

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    ws.auto_filter.ref = ws.dimensions

    # Column widths
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # Format percentages
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if isinstance(c.value, float) and 0 <= c.value <= 1:
                c.style = pct_style

    wb.save(output_file)
    print(f"✅ Excel file saved to {output_file}")

# --- Main ---
def main():
    print("Long Call/Put Option Analysis Tool")
    ticker = input("Enter ticker (e.g., NVDA): ").strip().upper()
    if not ticker:
        print("Ticker required.")
        return

    tk = yf.Ticker(ticker)
    S0 = tk.info.get('regularMarketPrice') or tk.info.get('previousClose') or None
    if S0 is None:
        hist = tk.history(period="5d")
        S0 = float(hist['Close'].iloc[-1]) if not hist.empty else None
    if S0 is None:
        print("Could not determine current price")
        return
    print(f"Current price: {S0}")

    exps = tk.options
    if not exps:
        print("No option expirations found")
        return
    print("\nAvailable expirations:")
    for i, e in enumerate(exps):
        print(f"[{i}] {e}")
    idx = int(input("Choose expiration index: "))
    expiry = dt.datetime.strptime(exps[idx], "%Y-%m-%d").date()

    move_input = input("Expected move percent (default 5%): ").strip()
    expected_move_pct = float(move_input)/100 if move_input else 0.05

    opt = tk.option_chain(exps[idx])
    calls, puts = opt.calls.copy(), opt.puts.copy()

    required_cols = ['strike','lastPrice','bid','ask','impliedVol']
    for col in required_cols:
        for df in [calls, puts]:
            if col not in df.columns:
                df[col] = float('nan')

    today = dt.date.today()
    calls_df = build_long_options('calls', calls, S0, expiry, today, expected_move_pct)
    puts_df = build_long_options('puts', puts, S0, expiry, today, expected_move_pct)

    all_options = pd.concat([calls_df, puts_df], ignore_index=True)
    all_options = all_options.sort_values(by=['return_ratio','expected_profit'], ascending=[False, False]).reset_index(drop=True)

    output_file = f"long_options_{ticker}_{expiry.strftime('%Y%m%d')}.xlsx"
    export_to_excel(all_options, output_file)

    print("\nTop 10 options by return ratio:")
    print(all_options[['option_type','strike','debit','expected_profit','return_ratio','prob_ITM']].head(10))

if __name__ == '__main__':
    main()
