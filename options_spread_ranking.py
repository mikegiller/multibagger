#!/usr/bin/env python3
"""
options_spread_ranking_fast.py

Fast version of options spread analyzer — filters for options traded today,
asks for expected move %, ranks vertical spreads, and exports results to Excel.
"""

import sys
import math
import datetime as dt
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import yfinance as yf
except ImportError:
    print("Install yfinance first: pip install yfinance")
    sys.exit(1)


# --- Helpers ---

def norm_cdf(x):
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def black_scholes_d1_d2(S, K, r, sigma, T):
    if sigma <= 0 or T <= 0:
        return None, None
    d1 = (math.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)
    return d1, d2


def option_mid_price(row):
    bid, ask, last = row.get("bid"), row.get("ask"), row.get("lastPrice")
    if bid is None or ask is None or (bid == 0 and ask == 0):
        return last or 0.0
    return (bid + ask) / 2.0


def build_spreads(option_type, chain_df, S0, expiry_date, today, r=0.0,
                  expected_move_pct=0.05, max_width=10):
    T_days = (expiry_date - today).days
    if T_days <= 0:
        return pd.DataFrame()
    T = T_days / 365.0

    strikes = sorted(chain_df['strike'].unique())
    rows = []

    for i, K1 in enumerate(strikes):
        for j in range(i + 1, min(i + 1 + max_width, len(strikes))):
            K2 = strikes[j]
            if option_type == 'calls':
                long_row = chain_df[chain_df['strike'] == K1].iloc[0].to_dict()
                short_row = chain_df[chain_df['strike'] == K2].iloc[0].to_dict()
            else:
                long_row = chain_df[chain_df['strike'] == K2].iloc[0].to_dict()
                short_row = chain_df[chain_df['strike'] == K1].iloc[0].to_dict()

            long_price = option_mid_price(long_row)
            short_price = option_mid_price(short_row)
            debit = long_price - short_price
            width = abs(K2 - K1)
            max_profit = width - debit
            max_loss = debit

            ST_expected = S0 * (1 + expected_move_pct) if option_type == 'calls' else S0 * (1 - expected_move_pct)
            if option_type == 'calls':
                profit_at_ST = max(0, ST_expected - K1) - max(0, ST_expected - K2) - debit
            else:
                profit_at_ST = max(0, K2 - ST_expected) - max(0, K1 - ST_expected) - debit

            return_ratio = profit_at_ST / debit if debit != 0 else 0.0
            iv = long_row.get('impliedVol') or short_row.get('impliedVol') or 0.0

            try:
                d1, d2 = black_scholes_d1_d2(S0, long_row['strike'], r, iv, T)
                prob_ITM = 1 - norm_cdf(d2) if option_type == 'calls' else norm_cdf(-d2)
            except:
                prob_ITM = None

            rows.append({
                'option_type': option_type,
                'long_strike': K1 if option_type == 'calls' else K2,
                'short_strike': K2 if option_type == 'calls' else K1,
                'width': width,
                'long_mid': long_price,
                'short_mid': short_price,
                'debit': debit,
                'max_profit': max_profit,
                'max_loss': max_loss,
                'profit_at_expected_ST': profit_at_ST,
                'return_ratio': return_ratio,
                'long_iv': iv,
                'prob_long_ITM': prob_ITM,
                'expiry': expiry_date.strftime("%Y-%m-%d"),
                'T_days': T_days
            })

    return pd.DataFrame(rows)

def export_to_excel(calls, puts, ranked_spreads, output_file):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    # Remove timezone info from datetime columns
    for df in [calls, puts, ranked_spreads]:
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.tz_localize(None)

    # Round numeric columns
    for df in [calls, puts, ranked_spreads]:
        df = df.round(2)
    
    # Write to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        calls.to_excel(writer, index=False, sheet_name='Calls')
        puts.to_excel(writer, index=False, sheet_name='Puts')
        ranked_spreads.to_excel(writer, index=False, sheet_name='Ranked Spreads')

        writer.sheets['Calls'].auto_filter.ref = writer.sheets['Calls'].dimensions
        writer.sheets['Puts'].auto_filter.ref = writer.sheets['Puts'].dimensions
        writer.sheets['Ranked Spreads'].auto_filter.ref = writer.sheets['Ranked Spreads'].dimensions

    # Load workbook for styling
    wb = load_workbook(output_file)

    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    bold_font = Font(bold=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Style header row
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center")
        
        # Shade rows for "Puts"
        if sheet_name == "Puts":
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = gray_fill

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(10, min(max_length + 2, 40))

    wb.save(output_file)



def main():
    print("Fast Options Spread Ranking Tool")
    ticker = input("Enter ticker (e.g., NVDA): ").strip().upper()
    if not ticker:
        return

    tk = yf.Ticker(ticker)

    # Faster price fetch
    hist = tk.history(period="1d")
    if hist.empty:
        print("Could not fetch price data.")
        return
    S0 = float(hist['Close'].iloc[-1])
    print(f"Current {ticker} price: {S0:.2f}")

    exps = tk.options
    if not exps:
        print("No expirations available.")
        return
    for i, e in enumerate(exps):
        print(f"[{i}] {e}")
    idx = int(input("Choose expiration index: "))
    expiry = dt.datetime.strptime(exps[idx], "%Y-%m-%d").date()

    # ✅ Expected move percent prompt (restored)
    move_input = input("Expected move percent (default 5): ").strip()
    expected_move_pct = float(move_input) / 100 if move_input else 0.05
    print(f"Using expected move: {expected_move_pct*100:.1f}%")

    chain = tk.option_chain(exps[idx])
    calls, puts = chain.calls.copy(), chain.puts.copy()

    today = datetime.now().date()
    if "lastTradeDate" in calls.columns:
        calls = calls[calls['lastTradeDate'].dt.date == today]
    if "lastTradeDate" in puts.columns:
        puts = puts[puts['lastTradeDate'].dt.date == today]
    print(f"Filtered to {len(calls)} calls and {len(puts)} puts traded today ({today}).")

    req_cols = ['contractSymbol', 'strike', 'lastPrice', 'bid', 'ask', 'impliedVol']
    for col in req_cols:
        for df in [calls, puts]:
            if col not in df.columns:
                df[col] = float('nan')

    call_spreads = build_spreads('calls', calls, S0, expiry, today, expected_move_pct=expected_move_pct)
    put_spreads = build_spreads('puts', puts, S0, expiry, today, expected_move_pct=expected_move_pct)

    all_spreads = pd.concat([call_spreads, put_spreads], ignore_index=True)
    all_spreads['spread_name'] = all_spreads.apply(
        lambda r: f"{r['option_type'][0].upper()} {r['long_strike']}/{r['short_strike']}", axis=1)
    ranked_spreads = all_spreads.sort_values(by='return_ratio', ascending=False).reset_index(drop=True)

    output_file = f"spread_rankings_{ticker}_{expiry.strftime('%Y%m%d')}_FAST.xlsx"
    export_to_excel(calls, puts, ranked_spreads, output_file)

    print("\nTop 10 spreads:")
    print(ranked_spreads[['spread_name', 'option_type', 'debit', 'profit_at_expected_ST', 'return_ratio']].head(10))


if __name__ == "__main__":
    main()
